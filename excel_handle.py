#!/usr/bin/env python
# encoding: utf-8
# Date: 2020/11/26
# file: copytest.py
# Email:
# Author: 唐政

from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font


def lt_month_handle(src_file, sheet_name, excel_list):
    #        src_file是源xlsx文件，
    #        sheet_name是目标xlsx里的新sheet名称

    src_workbook = load_workbook(src_file)

    # 模板 sheet
    sheet = src_workbook['1']
    sheet.title = sheet_name

    s_l_count = 0
    p_l_count = 0
    c_l_count = 0
    mn_l_count = 0
    z_l_count = len(excel_list)

    min_row = 4
    max_row = min_row + z_l_count
    max_column = 8

    remark_str = """备注：不合格描述  
    1.红色代表S不合格{}炉    
    2.蓝色代表P不合格{}炉
    3.绿色代表C不合格{}炉
    4.紫色代表Mn不合格{}炉"""
    red = Font(name='宋体', charset=134, color="FF0000")
    blue = Font(name='宋体', charset=134, color="00C070")
    green = Font(name='宋体', charset=134, color="00B050")
    purple = Font(name='宋体', charset=134, color="800080")

    # 表头
    sheet['A1'].value = '{}月份不合格钢水'.format(sheet_name)

    # 内容
    for m in range(min_row, max_row):
        # 设置行高如第四行
        sheet.row_dimensions[m].height = sheet.row_dimensions[min_row].height

        for n in range(1, max_column + 1):
            c = get_column_letter(n)

            # 单元格编号
            i = '%s%d' % (c, m)
            j = '%s%d' % (c, min_row)

            try:
                print("cell(%s)" % sheet[i])
                getattr(sheet.cell(row=m, column=n), "value")
                sheet[i].border = copy(sheet[j].border)
                sheet[i].fill = copy(sheet[j].fill)
                sheet[i].number_format = copy(sheet[j].number_format)
                sheet[i].protection = copy(sheet[j].protection)
                sheet[i].alignment = copy(sheet[j].alignment)

                # 赋值 判定结果颜色设置
                if n != max_column:
                    sheet[i].font = copy(sheet[j].font)
                    sheet[i].value = excel_list[m-min_row][n-1]
                else:
                    decide = []
                    el_dict_list = excel_list[m - min_row][n - 1]
                    for item in el_dict_list:
                        decide.append('{}:{}'.format(item['el'], item['value']))

                    if decide:
                        sheet[i].value = '{}超标判不合格F'.format(' '.join(decide))
                        if el_dict_list[0]['el'] == 'C':
                            sheet[i].font = green
                            c_l_count += 1
                        elif el_dict_list[0]['el'] == 'S':
                            sheet[i].font = red
                            s_l_count += 1
                        elif el_dict_list[0]['el'] == 'p':
                            sheet[i].font = blue
                            p_l_count += 1
                        elif el_dict_list[0]['el'] == 'Mn':
                            sheet[i].font = purple
                            mn_l_count += 1
                        else:
                            pass
                    else:
                        sheet[i].value = '{}需要重新判定'.format(excel_list[m-min_row])
            except AttributeError as e:
                print("cell({}) is {}".format(i, e))
                continue

    # 备注合并
    start_row = max_row + 1
    end_row = max_row + 4
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=max_column)

    # 备注
    b = 'A%d' % (max_row + 1)
    sheet[b].value = remark_str.format(s_l_count, p_l_count, c_l_count, mn_l_count)
    sheet[b].alignment = Alignment(wrapText=True)

    # 总计
    c = get_column_letter(max_column)
    b = '%s%d' % (c, end_row + 1)
    sheet[b].value = "共计:{}炉".format(z_l_count)
    sheet[b].alignment = Alignment(horizontal='center', vertical='center')

    src_workbook.save('test.xlsx')
    src_workbook.close()


