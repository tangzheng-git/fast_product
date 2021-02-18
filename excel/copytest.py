#!/usr/bin/env python
# encoding: utf-8
# Date: 2020/11/26
# file: copytest.py
# Email:
# Author: 唐政 

from copy import copy
from openpyxl import load_workbook, Workbook


def copy_xlsx(src_file, tag_file, sheet_name):
    #        src_file是源xlsx文件，
    #        tag_file是目标xlsx文件，
    #        sheet_name是目标xlsx里的新sheet名称

    src_workbook = load_workbook(src_file)

    try:
        tag_workbook = load_workbook(tag_file)
    except FileNotFoundError:
        tag_workbook = Workbook()

    # 模板 sheet
    default_sheet = src_workbook['1']

    # 目标 sheet 存在移除
    print(tag_workbook.sheetnames)
    if sheet_name in tag_workbook.sheetnames:
        sheet = tag_workbook[sheet_name]
        tag_workbook.remove(sheet)
        new_sheet = tag_workbook.create_sheet(sheet_name)
    else:
        new_sheet = tag_workbook.create_sheet(sheet_name)

    max_row = default_sheet.max_row  # 最大行数
    max_column = default_sheet.max_column  # 最大列数

    # 开始处理合并单元格
    merge_cell_list = list(default_sheet.merged_cells)
    if len(merge_cell_list) > 0:
        for i in range(0, len(merge_cell_list)):
            merge_cell = str(merge_cell_list[i]).replace('(<MergeCell ', '').replace('>,)', '')
            print("合并单元格 : %s" % merge_cell)
            new_sheet.merge_cells(merge_cell)

    for m in range(1, max_row + 1):
        new_sheet.row_dimensions[m].height = default_sheet.row_dimensions[m].height
        for n in range(1, 1 + max_column):

            if n < 27:
                c = chr(n + 64).upper()
            else:
                if n < 677:
                    c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
                else:
                    c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(
                        divmod(divmod(n, 676)[1], 26)[1] + 64)
            # 单元格编号
            i = '%s%d' % (c, m)

            if m == 1:
                # print("Modify column %s width from %d to %d" % (i, new_sheet.column_dimensions[c].width,
                #                                                 default_sheet.column_dimensions[c].width))
                new_sheet.column_dimensions[c].width = default_sheet.column_dimensions[c].width
            try:
                getattr(default_sheet.cell(row=m, column=n), "value")

                # 赋值到new_sheet单元格
                new_sheet[i].value = default_sheet[i].value

                # 拷贝格式
                if default_sheet[i].has_style:
                    new_sheet[i].font = copy(default_sheet[i].font)
                    new_sheet[i].border = copy(default_sheet[i].border)
                    new_sheet[i].fill = copy(default_sheet[i].fill)
                    new_sheet[i].number_format = copy(default_sheet[i].number_format)
                    new_sheet[i].protection = copy(default_sheet[i].protection)
                    new_sheet[i].alignment = copy(default_sheet[i].alignment)
            except AttributeError as e:
                # print("cell(%s) is %s" % (i, e))
                continue

    tag_workbook.save(tag_file)

    src_workbook.close()
    tag_workbook.close()


month = 11
src_file = '不合格钢水统计表.xlsx'
tag_file = 'test.xlsx'
sheet_name = str(month)

copy_xlsx(src_file, tag_file, sheet_name)
