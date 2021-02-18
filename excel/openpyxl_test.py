#!/usr/bin/env python
# encoding: utf-8
# Date: 2020/11/26
# file: openpyxl_test.py
# Email:
# Author: 唐政 

from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Color
from openpyxl.styles import PatternFill, \
    Border, Side, Alignment, Protection, Font



def deal_excel_by_openpyxl(src_file, tag_file, sheet_name):
    workbook = load_workbook(src_file)
    worksheet = workbook['1']

    worksheet.title = sheet_name
    side = Side(border_style='thin')

    align = Alignment()

    font = Font(name='宋体', charset=134, family=None, bold=None,
        italic=None, strikethrough=None, outline=None, shadow=None,
        condense=None, color='00ff00', extend=None, u=None, vertAlign=None,
        scheme=None)

    worksheet['A6'].value = '123456789123ha哈哈哈'
    worksheet['A6'].font = font

    workbook.save(tag_file)
    workbook.close()


src_file = '不合格钢水统计表.xlsx'
tag_file = 'test.xlsx'
sheet_name = '测试'

deal_excel_by_openpyxl(src_file, tag_file, sheet_name)
